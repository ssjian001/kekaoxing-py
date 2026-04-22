"""甘特图主视图 - 左侧任务列表 + 右侧甘特条时间线"""

from __future__ import annotations
from datetime import date, timedelta

from PySide6.QtWidgets import (
    QWidget, QHBoxLayout, QVBoxLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QSplitter, QLabel, QScrollArea,
    QMenu, QMessageBox, QInputDialog, QLineEdit, QComboBox, QSlider,
    QDialog, QPushButton, QToolTip,
)
from PySide6.QtCore import Qt, QRectF, QPoint, QPointF, Signal, QTimer
from PySide6.QtGui import (
    QPainter, QColor, QBrush, QPen, QFont, QFontMetrics,
    QLinearGradient, QPainterPath,
   QCursor,
)

from src.db.database import Database
from src.models import (
    Task, Section, DEFAULT_SECTION_LABELS, DEFAULT_SECTION_COLORS,
    get_section_label, get_section_color,
    ScheduleConfig, ScheduleMode,
)
from src.core.scheduler import run_auto_schedule


# ── 常量 ───────────────────────────────────────────

DAY_WIDTH = 40
ROW_HEIGHT = 36
HEADER_HEIGHT = 40
LEFT_MARGIN = 8
BAR_RADIUS = 6

COL_NUM = 0
COL_NAME = 1
COL_SECTION = 2
COL_DURATION = 3
COL_START = 4
COL_PROGRESS = 5
COL_COLS = 6


# ── 甘特条绘制区 ─────────────────────────────────

class GanttCanvas(QWidget):
    """自绘甘特图区域，包含任务条和依赖连线"""

    task_moved = Signal(int, int)       # task_id, new_day（拖拽释放后）
    task_selected = Signal(int)
    task_double_clicked = Signal(int)
    task_right_clicked = Signal(int, QPoint)
    zoom_changed = Signal(int)          # 新的 day_width 值
    task_resized = Signal(int, int)     # task_id, new_duration（右边缘拖拽释放后）

    def __init__(self, parent=None):
        super().__init__(parent)
        self.tasks: list[Task] = []
        self.total_days = 30
        self.start_date = date.today()
        self.selected_task_id = -1
        self.day_width = DAY_WIDTH       # 可缩放
        self._dragging = False
        self._drag_task_id = -1
        self._drag_original_day = 0      # 拖拽前的 start_day（用于确认/取消）
        self._drag_preview_day = 0       # 拖拽中的预览 day
        self._drag_offset = 0
        self._hover_task_id = -1         # 悬停任务（用于 tooltip）
        self._drag_mode = "move"          # "move" 或 "resize"
        self._drag_original_duration = 1 # resize 前的 duration
        self._drag_preview_duration = 1  # resize 中的预览 duration
        self.setMouseTracking(True)
        self.setMinimumWidth(600)

    def set_tasks(self, tasks: list[Task], total_days: int = 30):
        self.tasks = tasks
        self.total_days = max(total_days, 30)
        dw = self.day_width
        self.setMinimumHeight(HEADER_HEIGHT + len(tasks) * ROW_HEIGHT + 20)
        self.setMinimumWidth(LEFT_MARGIN + self.total_days * dw + 20)
        self.update()

    def set_start_date(self, d: date):
        self.start_date = d
        self.update()

    def date_for_day(self, day: int) -> date:
        return self.start_date + timedelta(days=day)

    def _task_y(self, task: Task) -> int:
        idx = next((i for i, t in enumerate(self.tasks) if t.id == task.id), -1)
        if idx < 0:
            return -1
        return HEADER_HEIGHT + idx * ROW_HEIGHT

    # ── 绘制 ────────────────────────────────────

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        w = self.width()
        h = self.height()
        dw = self.day_width

        # 背景
        p.fillRect(0, 0, w, h, QColor("#11111b"))

        # 空状态引导
        if not self.tasks:
            p.setPen(QColor("#a6adc8"))
            p.setFont(QFont("sans-serif", 14))
            p.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter,
                "📋 暂无测试任务\n\n点击「➕ 添加任务」或「📋 任务模板」开始\n快捷键: Ctrl+T")
            p.end()
            return

        # 日期头
        p.fillRect(0, 0, w, HEADER_HEIGHT, QColor("#181825"))
        font_header = QFont("Microsoft YaHei", 9)
        font_header.setBold(True)
        p.setFont(font_header)
        p.setPen(QColor("#a6adc8"))

        today = date.today()
        today_day = (today - self.start_date).days if today >= self.start_date else -999

        for d in range(self.total_days + 1):
            x = LEFT_MARGIN + d * dw
            if x > w:
                break
            dt = self.date_for_day(d)
            dow = dt.weekday()
            if dow >= 5:
                p.fillRect(x, HEADER_HEIGHT, dw, h - HEADER_HEIGHT, QColor("#18182580"))
            if d == today_day:
                p.fillRect(x, 0, dw, h, QColor("#89b4fa18"))
            p.setPen(QPen(QColor("#313244"), 1))
            p.drawLine(x, HEADER_HEIGHT, x, h)
            p.setPen(QColor("#89b4fa") if d == today_day else QColor("#a6adc8"))
            label = f"{dt.month}/{dt.day}"
            wd = ["一", "二", "三", "四", "五", "六", "日"][dow]
            p.drawText(QRectF(x + 2, 2, dw - 4, 18), Qt.AlignCenter, label)
            p.drawText(QRectF(x + 2, 18, dw - 4, 18), Qt.AlignCenter, f"周{wd}")

        # 依赖连线（在任务条下面画）
        self._draw_dependencies(p)

        # 任务行
        p.setFont(QFont("Microsoft YaHei", 10))
        for i, task in enumerate(self.tasks):
            y = HEADER_HEIGHT + i * ROW_HEIGHT
            rect = QRectF(0, y, w, ROW_HEIGHT)
            if i % 2 == 0:
                p.fillRect(rect, QColor("#1e1e2e08"))
            if task.id == self.selected_task_id:
                p.fillRect(rect, QColor("#89b4fa15"))
            p.setPen(QPen(QColor("#31324440"), 1))
            p.drawLine(0, y + ROW_HEIGHT, w, y + ROW_HEIGHT)
            self._draw_bar(p, task, y)

        # 今天竖线
        if 0 <= today_day <= self.total_days:
            x = LEFT_MARGIN + today_day * self.day_width + self.day_width / 2
            p.setPen(QPen(QColor("#f38ba8"), 2, Qt.DashLine))
            p.drawLine(int(x), HEADER_HEIGHT, int(x), h)

        # 拖拽预览条（半透明）
        if self._dragging and self._drag_task_id > 0:
            drag_task = next((t for t in self.tasks if t.id == self._drag_task_id), None)
            if drag_task:
                idx = next((i for i, t in enumerate(self.tasks) if t.id == self._drag_task_id), -1)
                if idx >= 0:
                    preview_y = HEADER_HEIGHT + idx * ROW_HEIGHT + 5
                    if self._drag_mode == "resize":
                        preview_x = LEFT_MARGIN + drag_task.start_day * self.day_width + 2
                        preview_w = max(self._drag_preview_duration * self.day_width - 4, 20)
                    else:
                        preview_x = LEFT_MARGIN + self._drag_preview_day * self.day_width + 2
                        preview_w = max(drag_task.duration * self.day_width - 4, 20)
                    p.setPen(Qt.NoPen)
                    p.setBrush(QBrush(QColor("#89b4fa50")))
                    p.drawRoundedRect(QRectF(preview_x, preview_y, preview_w, ROW_HEIGHT - 10), BAR_RADIUS, BAR_RADIUS)
                    # 预览 day 标签
                    p.setPen(QPen(QColor("#89b4fa"), 1))
                    p.setFont(QFont("Microsoft YaHei", 8))
                    if self._drag_mode == "resize":
                        p.drawText(QRectF(preview_x, preview_y - 16, preview_w, 14),
                                   Qt.AlignCenter, f"{self._drag_preview_duration}天")
                    else:
                        p.drawText(QRectF(preview_x, preview_y - 16, preview_w, 14),
                                   Qt.AlignCenter, f"D{self._drag_preview_day + 1}")

        p.end()

    def _draw_bar(self, p: QPainter, task: Task, y: int):
        """绘制单个任务条"""
        dw = self.day_width
        bar_x = LEFT_MARGIN + task.start_day * dw + 2
        bar_w = max(task.duration * dw - 4, 20)
        bar_y = y + 5
        bar_h = ROW_HEIGHT - 10

        color = QColor(get_section_color(task.section, getattr(self, '_section_colors', None)))
        if task.done:
            color = color.darker(150)

        grad = QLinearGradient(bar_x, bar_y, bar_x, bar_y + bar_h)
        grad.setColorAt(0, color.lighter(120))
        grad.setColorAt(1, color)
        p.setBrush(QBrush(grad))

        if task.id == self.selected_task_id:
            p.setPen(QPen(QColor("#89b4fa"), 2))
        else:
            p.setPen(QPen(color.darker(140), 1))

        p.drawRoundedRect(QRectF(bar_x, bar_y, bar_w, bar_h), BAR_RADIUS, BAR_RADIUS)

        # 进度条
        if task.progress > 0:
            prog_w = bar_w * (task.progress / 100.0)
            prog_rect = QRectF(bar_x, bar_y + bar_h - 4, prog_w, 4)
            p.setBrush(QBrush(QColor("#a6e3a1")))
            p.setPen(Qt.NoPen)
            p.drawRoundedRect(prog_rect, 2, 2)

        # 完成勾
        if task.done:
            p.setPen(QPen(QColor("#a6e3a1"), 3))
            cx = bar_x + bar_w / 2
            cy = bar_y + bar_h / 2
            p.drawLine(int(cx - 6), int(cy), int(cx - 2), int(cy + 5))
            p.drawLine(int(cx - 2), int(cy + 5), int(cx + 6), int(cy - 4))

        # 文字
        text_color = QColor("#1e1e2e") if color.lightness() > 150 else QColor("#cdd6f4")
        p.setPen(text_color)
        text = f"{task.num} {task.name_cn}"
        fm = QFontMetrics(p.font())
        if fm.horizontalAdvance(text) > bar_w - 8:
            text = f"{task.num} {task.name_cn[:4]}"
        if not task.done:
            p.drawText(QRectF(bar_x + 6, bar_y, bar_w - 12, bar_h),
                       Qt.AlignVCenter | Qt.AlignLeft, text)

        # 标签色点
        tag_defs = getattr(self, '_tag_defs', {})
        task_tags = getattr(self, '_task_tags_map', {})
        if task.id in task_tags:
            for i, tid in enumerate(task_tags[task.id][:3]):
                td = tag_defs.get(tid)
                if td:
                    dot_x = bar_x + bar_w - 8 - i * 10
                    dot_y = bar_y + 8
                    p.setPen(Qt.NoPen)
                    p.setBrush(QBrush(QColor(td.get("color", "#89b4fa"))))
                    p.drawEllipse(QPointF(dot_x, dot_y), 4, 4)

    def _draw_dependencies(self, p: QPainter):
        """绘制依赖关系连线（箭头）"""
        task_map = {t.id: t for t in self.tasks}
        pen = QPen(QColor("#f9e2af88"), 2)
        p.setPen(pen)
        p.setBrush(QBrush(QColor("#f9e2af88")))

        for task in self.tasks:
            for dep_num in task.dependencies:
                dep = next((t for t in self.tasks if t.num == dep_num), None)
                if not dep:
                    continue

                # 起点线：依赖任务条的右端中点
                src_x = LEFT_MARGIN + (dep.start_day + dep.duration) * self.day_width
                src_idx = next((i for i, t in enumerate(self.tasks) if t.id == dep.id), -1)
                if src_idx < 0:
                    continue
                src_y = HEADER_HEIGHT + src_idx * ROW_HEIGHT + ROW_HEIGHT / 2

                # 终点线：当前任务条的左端中点
                dst_x = LEFT_MARGIN + task.start_day * self.day_width + 2
                dst_idx = next((i for i, t in enumerate(self.tasks) if t.id == task.id), -1)
                if dst_idx < 0:
                    continue
                dst_y = HEADER_HEIGHT + dst_idx * ROW_HEIGHT + ROW_HEIGHT / 2

                # 贝塞尔曲线
                mid_x = (src_x + dst_x) / 2
                path = QPainterPath()
                path.moveTo(src_x, src_y)
                path.cubicTo(mid_x, src_y, mid_x, dst_y, dst_x, dst_y)
                p.drawPath(path)

                # 箭头
                arrow_size = 6
                p.setBrush(QBrush(QColor("#f9e2af")))
                triangle = QPainterPath()
                triangle.moveTo(dst_x, dst_y)
                triangle.lineTo(dst_x - arrow_size, dst_y - arrow_size / 2)
                triangle.lineTo(dst_x - arrow_size, dst_y + arrow_size / 2)
                triangle.closeSubpath()
                p.drawPath(triangle)
                p.setBrush(QBrush(QColor("#f9e2af88")))

    # ── 鼠标交互 ────────────────────────────────

    def _task_at_y(self, y: int) -> Task | None:
        if y < HEADER_HEIGHT:
            return None
        row = (y - HEADER_HEIGHT) // ROW_HEIGHT
        if 0 <= row < len(self.tasks):
            return self.tasks[row]
        return None

    def _bar_rect_for_task(self, task: Task) -> tuple:
        """返回任务条的 (bar_x, bar_y, bar_w, bar_h)"""
        dw = self.day_width
        idx = next((i for i, t in enumerate(self.tasks) if t.id == task.id), -1)
        if idx < 0:
            return (0, 0, 0, 0)
        bar_x = LEFT_MARGIN + task.start_day * dw + 2
        bar_w = max(task.duration * dw - 4, 20)
        bar_y = HEADER_HEIGHT + idx * ROW_HEIGHT + 5
        bar_h = ROW_HEIGHT - 10
        return (bar_x, bar_y, bar_w, bar_h)

    def _is_on_right_edge(self, task: Task, mouse_x: int) -> bool:
        """判断鼠标是否在任务条右边缘（±5px）"""
        bar_x, bar_y, bar_w, bar_h = self._bar_rect_for_task(task)
        right_edge = bar_x + bar_w
        return abs(mouse_x - right_edge) <= 5

    def _day_at_x(self, x: int) -> int:
        """将 x 坐标吸附到最近的 day"""
        dw = self.day_width
        day = round((x - LEFT_MARGIN) / dw)
        return max(0, day)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            task = self._task_at_y(event.pos().y())
            if task:
                self.selected_task_id = task.id
                self.task_selected.emit(task.id)
                self._dragging = True
                self._drag_task_id = task.id
                if self._is_on_right_edge(task, event.pos().x()):
                    self._drag_mode = "resize"
                    self._drag_original_duration = task.duration
                    self._drag_preview_duration = task.duration
                else:
                    self._drag_mode = "move"
                    self._drag_original_day = task.start_day
                    self._drag_preview_day = task.start_day
                    self._drag_offset = event.pos().x() - (LEFT_MARGIN + task.start_day * self.day_width)
                self.update()
            else:
                self.selected_task_id = -1
                self.update()

    def mouseMoveEvent(self, event):
        if self._dragging:
            task = next((t for t in self.tasks if t.id == self._drag_task_id), None)
            if not task:
                return

            if self._drag_mode == "resize":
                # resize 模式：调整工期
                bar_x, _, _, _ = self._bar_rect_for_task(task)
                new_duration = max(1, round((event.pos().x() - bar_x) / self.day_width))
                if new_duration != self._drag_preview_duration:
                    self._drag_preview_duration = new_duration
                    task.duration = new_duration
                    self.update()
                # tooltip
                dt = self.start_date + timedelta(days=task.start_day)
                end_dt = dt + timedelta(days=new_duration)
                QToolTip.showText(
                    event.globalPosition().toPoint(),
                    f"📋 {task.num} {task.name_cn}\n"
                    f"📏 工期: {new_duration} 天\n"
                    f"📅 {dt.strftime('%m/%d')} → {end_dt.strftime('%m/%d')}",
                    self,
                )
            else:
                # move 模式：拖拽移动
                day = self._day_at_x(event.pos().x() - self._drag_offset)
                day = max(0, day)
                if day != self._drag_preview_day:
                    self._drag_preview_day = day
                    self.update()
                # 显示 tooltip
                dt = self.start_date + timedelta(days=day)
                end_dt = dt + timedelta(days=task.duration)
                QToolTip.showText(
                    event.globalPosition().toPoint(),
                    f"📋 {task.num} {task.name_cn}\n"
                    f"📍 D{day + 1} → D{day + task.duration}\n"
                    f"📅 {dt.strftime('%m/%d')} → {end_dt.strftime('%m/%d')}",
                    self,
                )
        else:
            # 悬停 tooltip
            task = self._task_at_y(event.pos().y())
            if task and task.id != self._hover_task_id:
                self._hover_task_id = task.id
                # 设置光标：右边缘用 SizeHorCursor，内部用 OpenHandCursor
                if self._is_on_right_edge(task, event.pos().x()):
                    self.setCursor(QCursor(Qt.CursorShape.SizeHorCursor))
                else:
                    self.setCursor(QCursor(Qt.CursorShape.OpenHandCursor))

                dt = self.start_date + timedelta(days=task.start_day)
                end_dt = dt + timedelta(days=task.duration)
                QToolTip.showText(
                    event.globalPosition().toPoint(),
                    f"📋 {task.num} {task.name_cn}\n"
                    f"📅 {dt.strftime('%m/%d')} → {end_dt.strftime('%m/%d')} ({task.duration}天)\n"
                    f"📊 进度: {task.progress:.0f}%  {'✅ 已完成' if task.done else '⏳ 进行中'}",
                    self,
                )
            elif not task and self._hover_task_id >= 0:
                self._hover_task_id = -1
                self.setCursor(QCursor(Qt.CursorShape.ArrowCursor))
                QToolTip.hideText()

    def mouseReleaseEvent(self, event):
        if self._dragging and self._drag_task_id > 0:
            if self._drag_mode == "resize":
                new_duration = self._drag_preview_duration
                if new_duration != self._drag_original_duration:
                    self.task_resized.emit(self._drag_task_id, new_duration)
                else:
                    # 恢复原值
                    task = next((t for t in self.tasks if t.id == self._drag_task_id), None)
                    if task:
                        task.duration = self._drag_original_duration
                        self.update()
            else:
                new_day = self._drag_preview_day
                # 只有实际移动了才发射信号
                if new_day != self._drag_original_day:
                    self.task_moved.emit(self._drag_task_id, new_day)
        self._dragging = False
        self._drag_task_id = -1
        self._drag_mode = "move"
        self.setCursor(QCursor(Qt.CursorShape.ArrowCursor))
        QToolTip.hideText()

    def wheelEvent(self, event):
        """Ctrl+滚轮缩放"""
        if event.modifiers() & Qt.ControlModifier:
            delta = event.angleDelta().y()
            if delta > 0:
                self.day_width = min(self.day_width + 4, 120)
            elif delta < 0:
                self.day_width = max(self.day_width - 4, 16)
            self.zoom_changed.emit(self.day_width)
            # 重新计算最小宽度
            self.setMinimumWidth(LEFT_MARGIN + self.total_days * self.day_width + 20)
            self.update()
            event.accept()
        else:
            super().wheelEvent(event)

    def mouseDoubleClickEvent(self, event):
        task = self._task_at_y(event.pos().y())
        if task:
            self.task_double_clicked.emit(task.id)

    def contextMenuEvent(self, event):
        task = self._task_at_y(event.pos().y())
        if task:
            self.selected_task_id = task.id
            self.task_selected.emit(task.id)
            self.task_right_clicked.emit(task.id, event.globalPos())


# ── 甘特图主视图 ──────────────────────────────────

class GanttView(QWidget):
    task_selected = Signal(int)  # 向外转发 GanttCanvas 的 task_selected 信号

    def __init__(self, db: Database, undo_manager=None, parent=None):
        super().__init__(parent)
        self.db = db
        self.undo_manager = undo_manager
        self.tasks: list[Task] = []
        self._filtered_tasks: list[Task] = []
        self._scheduler_dialog = None
        self._last_schedule_result = None
        self._collapsed_sections: set[str] = set()
        self._group_mode = True
        self._row_to_task: list[Task | None] = []  # table row → Task or None (group header)

        # 从数据库加载项目起始日期（toolbar 需要）
        sd_str = self.db.get_setting("start_date", date.today().strftime("%Y-%m-%d"))
        self._project_start_date = date.fromisoformat(sd_str)

        self._setup_ui()

        self.gantt_canvas.set_start_date(self._project_start_date)

        self.refresh()

    def set_project_start_date(self, d: date):
        """更新项目起始日期并持久化"""
        self._project_start_date = d
        self.db.set_setting("start_date", d.isoformat())
        self.gantt_canvas.set_start_date(d)
        self.update()

    def _on_date_label_clicked(self, event):
        """点击日期标签弹出日期选择器"""
        current = self._project_start_date
        text, ok = QInputDialog.getText(
            self, "设置起始日期", "输入起始日期 (YYYY-MM-DD):",
            text=current.isoformat()
        )
        if ok:
            try:
                d = date.fromisoformat(text.strip())
                self.set_project_start_date(d)
                self._date_label.setText(f"📅 {d.isoformat()}")
                self.refresh()
            except ValueError:
                QMessageBox.warning(self, "格式错误", "请输入 YYYY-MM-DD 格式的日期")

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        # ── 搜索/过滤工具栏 ──
        self._setup_toolbar(layout)

        # ── 垂直 Splitter：上 = 甘特图区域，下 = 资源时间线 ──
        self._vsplitter = QSplitter(Qt.Vertical)
        layout.addWidget(self._vsplitter)

        # ── 上半部分：任务表格 + 甘特图（水平 Splitter，滚动同步）──
        h_splitter = QSplitter(Qt.Horizontal)

        # 左侧任务列表
        self.task_table = QTableWidget()
        self.task_table.setColumnCount(COL_COLS)
        self.task_table.setHorizontalHeaderLabels(
            ["编号", "测试项目", "分类", "天数", "开始", "进度%"]
        )
        header = self.task_table.horizontalHeader()
        # 编号列固定窄，分类/天数/开始/进度 ResizeToContents，测试项目拉伸
        header.setSectionResizeMode(COL_NUM, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(COL_NAME, QHeaderView.Stretch)
        header.setSectionResizeMode(COL_SECTION, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(COL_DURATION, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(COL_START, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(COL_PROGRESS, QHeaderView.ResizeToContents)
        self.task_table.verticalHeader().setVisible(False)
        self.task_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.task_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.task_table.setSelectionMode(QAbstractItemView.MultiSelection)
        self.task_table.setAlternatingRowColors(True)
        self.task_table.verticalHeader().setDefaultSectionSize(ROW_HEIGHT)

        # 组头行点击信号（折叠/展开）
        self.task_table.itemPressed.connect(self._on_table_item_clicked)

        # 右侧甘特图（带滚动区域）
        self.gantt_canvas = GanttCanvas()
        self.gantt_canvas.task_selected.connect(self._on_gantt_selected)
        # 向外转发选中信号
        self.gantt_canvas.task_selected.connect(self.task_selected)
        self.gantt_canvas.task_moved.connect(self._on_task_moved)
        self.gantt_canvas.task_double_clicked.connect(self._on_task_double_clicked)
        self.gantt_canvas.task_right_clicked.connect(self._show_context_menu)
        self.gantt_canvas.task_resized.connect(self._on_task_resized)

        self.gantt_scroll = QScrollArea()
        self.gantt_scroll.setWidget(self.gantt_canvas)
        self.gantt_scroll.setWidgetResizable(False)
        self.gantt_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        self.gantt_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.gantt_scroll.setFrameShape(QScrollArea.Shape.NoFrame)

        h_splitter.addWidget(self.task_table)
        h_splitter.addWidget(self.gantt_scroll)
        h_splitter.setStretchFactor(0, 2)
        h_splitter.setStretchFactor(1, 5)

        # 表格与画布垂直滚动同步
        self.task_table.verticalScrollBar().valueChanged.connect(
            self.gantt_scroll.verticalScrollBar().setValue
        )
        self.gantt_scroll.verticalScrollBar().valueChanged.connect(
            self.task_table.verticalScrollBar().setValue
        )

        # 选中行同步
        self.task_table.currentCellChanged.connect(
            lambda r, c, pr, pc: self._on_row_selected(r)
        )

        self._vsplitter.addWidget(h_splitter)

        # ── 下半部分：资源时间线（放在滚动区域内，可调大小）──
        from src.widgets.resource_timeline import ResourceTimeline

        timeline_container = QWidget()
        timeline_layout = QVBoxLayout(timeline_container)
        timeline_layout.setContentsMargins(0, 0, 0, 0)
        timeline_layout.setSpacing(0)

        self.resource_timeline_label = QLabel("  🔧 设备资源占用热力图")
        self.resource_timeline_label.setStyleSheet(
            "color: #a6adc8; font-weight: bold; padding: 8px 0;"
        )
        timeline_layout.addWidget(self.resource_timeline_label)

        # 资源时间线放进水平滚动区域，内容溢出时可滚动
        self.resource_timeline = ResourceTimeline()
        self.timeline_scroll = QScrollArea()
        self.timeline_scroll.setWidget(self.resource_timeline)
        self.timeline_scroll.setWidgetResizable(False)
        self.timeline_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.timeline_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.timeline_scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        timeline_layout.addWidget(self.timeline_scroll)

        self._vsplitter.addWidget(timeline_container)

        # 垂直分割比例：甘特图占 70%，资源时间线占 30%
        self._vsplitter.setStretchFactor(0, 7)
        self._vsplitter.setStretchFactor(1, 3)

        # 水平滚动同步：甘特图 ↔ 资源时间线
        self.gantt_scroll.horizontalScrollBar().valueChanged.connect(
            self._sync_timeline_scroll
        )

        # 缩放信号连接
        self.gantt_canvas.zoom_changed.connect(self._on_zoom_changed)

        # ── 键盘快捷键 ──
        from PySide6.QtGui import QShortcut, QKeySequence
        QShortcut(QKeySequence("Delete"), self, self._on_delete_selected)
        QShortcut(QKeySequence("Space"), self, self._on_toggle_done)
        QShortcut(QKeySequence("Ctrl+N"), self, self._add_new_task)

    def select_task_by_id(self, task_id: int):
        """根据 task_id 选中甘特图任务行并滚动到可见"""
        for row in range(self.task_table.rowCount()):
            item = self.task_table.item(row, 0)
            if item and item.data(Qt.ItemDataRole.UserRole) == task_id:
                self.task_table.selectRow(row)
                self.task_table.scrollTo(self.task_table.model().index(row, 0))
                break

    def refresh(self):
        # 同步起始日期到 Canvas
        sd_str = self.db.get_setting("start_date", date.today().strftime("%Y-%m-%d"))
        self._project_start_date = date.fromisoformat(sd_str)
        self.gantt_canvas.set_start_date(self._project_start_date)

        self.tasks = self.db.get_all_tasks()
        # 从数据库加载动态分类配置
        self._section_labels = self.db.get_section_labels()
        self._section_colors = self.db.get_section_colors()
        # 刷新分类过滤下拉框
        self._refresh_section_combo()
        self._apply_filter()
        self._update_timeline()
        # 更新统计面板
        bottleneck_count = 0
        if self._last_schedule_result:
            bottleneck_count = len(self._last_schedule_result.get("report", {}).get("bottlenecks", []))
        self.stats_panel.update_stats(self.tasks, self.db.get_all_sections(), bottleneck_count)
        # 加载标签数据
        self._load_tag_data()

    def _load_tag_data(self):
        """从数据库加载标签定义和任务-标签映射"""
        try:
            from src.widgets.task_tags import load_task_tags_from_db
            tag_definitions, task_tags = load_task_tags_from_db(self.db)
            self.gantt_canvas._tag_defs = {td["id"]: td for td in tag_definitions}
            self.gantt_canvas._task_tags_map = task_tags
        except Exception:
            self.gantt_canvas._tag_defs = {}
            self.gantt_canvas._task_tags_map = {}

    def _populate_table(self):
        tasks = self._filtered_tasks
        self._row_to_task = []
        self.task_table.setSortingEnabled(False)

        if not self._group_mode:
            # ── 扁平模式：与原来完全一致 ──
            self.task_table.setRowCount(len(tasks))
            for i, t in enumerate(tasks):
                self._row_to_task.append(t)
                self._set_task_row(i, t)
            return

        # ── 分组模式：按 section 分组 ──
        # 收集 section → 任务列表（保持出现顺序）
        section_order: list[str] = []
        section_tasks: dict[str, list[Task]] = {}
        for t in tasks:
            key = t.section.value if isinstance(t.section, Section) else str(t.section)
            if key not in section_tasks:
                section_order.append(key)
                section_tasks[key] = []
            section_tasks[key].append(t)

        # 计算总行数：每个分类 1 行组头 + (折叠时不加任务行)
        total_rows = 0
        for key in section_order:
            total_rows += 1  # 组头行
            if key not in self._collapsed_sections:
                total_rows += len(section_tasks[key])

        self.task_table.setRowCount(total_rows)
        row = 0
        labels = getattr(self, '_section_labels', None)

        for key in section_order:
            group_label = get_section_label(
                Section(key) if key in Section.__members__ else key, labels
            )
            count = len(section_tasks[key])
            collapsed = key in self._collapsed_sections
            icon = "▶ 🏷️" if collapsed else "▼ 🏷️"
            text = f"{icon} {group_label} ({count})"

            # 组头行
            header_item = QTableWidgetItem(text)
            header_item.setData(Qt.ItemDataRole.UserRole, ("group", key))
            header_item.setBackground(QColor("#181825"))
            font = header_item.font()
            font.setBold(True)
            header_item.setFont(font)
            header_item.setForeground(QColor("#cdd6f4"))
            header_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            # 设置组头行不可选中
            header_item.setFlags(header_item.flags() & ~Qt.ItemFlag.ItemIsSelectable)
            self.task_table.setItem(row, 0, header_item)
            # 其他列也放一个空 item，防止 span 后点击空白列异常
            for c in range(1, COL_COLS):
                dummy = QTableWidgetItem()
                dummy.setFlags(dummy.flags() & ~Qt.ItemFlag.ItemIsSelectable)
                self.task_table.setItem(row, c, dummy)
            self.task_table.setSpan(row, 0, 1, COL_COLS)
            self._row_to_task.append(None)  # 组头行没有关联任务
            row += 1

            if not collapsed:
                for t in section_tasks[key]:
                    self._row_to_task.append(t)
                    self._set_task_row(row, t)
                    row += 1

        self.task_table.setSortingEnabled(True)

    def _set_task_row(self, row: int, t: Task):
        """在表格指定行填充单个任务数据"""
        color = get_section_color(t.section, getattr(self, '_section_colors', None))
        items_data = [
            t.num,
            t.name_cn,
            get_section_label(t.section, getattr(self, '_section_labels', None)),
            str(t.duration),
            None,  # COL_START 占位，下面单独设置
            f"{t.progress:.0f}",
        ]
        for j, text in enumerate(items_data):
            item = QTableWidgetItem(text if text is not None else "")
            item.setData(Qt.ItemDataRole.UserRole, t.id)
            item.setTextAlignment(Qt.AlignCenter)
            if j == COL_NAME:
                item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            if j == COL_NUM:
                item.setForeground(QColor(color))
            if t.done:
                item.setForeground(QColor("#585b70"))
            self.task_table.setItem(row, j, item)

        # 开始天 → 显示日期
        start_item = self.task_table.item(row, COL_START)
        if start_item and hasattr(self, '_project_start_date'):
            dt = self._project_start_date + timedelta(days=t.start_day)
            start_item.setText(f"{dt.month}/{dt.day}")
            start_item.setToolTip(
                f"第{t.start_day}天 → {dt.isoformat()} "
                f"周{['一','二','三','四','五','六','日'][dt.weekday()]}"
            )

    def _on_table_item_clicked(self, item: QTableWidgetItem):
        """处理表格点击：组头行点击时切换折叠状态"""
        if item is None:
            return
        data = item.data(Qt.ItemDataRole.UserRole)
        if data and isinstance(data, tuple) and data[0] == "group":
            section_key = data[1]
            if section_key in self._collapsed_sections:
                self._collapsed_sections.discard(section_key)
            else:
                self._collapsed_sections.add(section_key)
            self._populate_table()
            # 甘特图不需要变化（_filtered_tasks 不变）
            self._update_gantt()

    def _on_group_toggle(self, checked: bool):
        """工具栏分组/扁平模式切换"""
        self._group_mode = checked
        if not checked:
            # 切到扁平模式时清空折叠状态
            self._collapsed_sections.clear()
        self._populate_table()

    def _update_gantt(self):
        max_day = max((t.start_day + t.duration for t in self._filtered_tasks), default=30)
        self.gantt_canvas.set_tasks(self._filtered_tasks, max_day)

    def _update_timeline(self):
        """根据当前排程结果更新资源时间线"""
        if not self._last_schedule_result:
            # 首次加载时自动跑一次排程获取时间线
            resources = self.db.get_all_resources()
            config = ScheduleConfig(mode=ScheduleMode.BALANCED)
            result = run_auto_schedule(self.tasks, resources, config)
            self._last_schedule_result = result

        self.resource_timeline.set_data(
            self._last_schedule_result.get("timeline", {}),
            self.db.get_all_resources(),
            self._last_schedule_result.get("report", {}).get("total_days", 30),
        )
        # 同步起始日期到资源时间线
        self.resource_timeline.set_start_date(self._project_start_date)

    def _on_row_selected(self, row: int):
        if 0 <= row < len(self._row_to_task):
            task = self._row_to_task[row]
            if task is not None:
                self.gantt_canvas.selected_task_id = task.id
                self.gantt_canvas.update()

    def _on_gantt_selected(self, task_id: int):
        for table_row, task in enumerate(self._row_to_task):
            if task is not None and task.id == task_id:
                self.task_table.selectRow(table_row)
                # 确保表格滚动到可见位置
                self.task_table.scrollTo(self.task_table.model().index(table_row, 0))
                break

    def _on_task_moved(self, task_id: int, new_day: int):
        for t in self.tasks:
            if t.id == task_id:
                old_day = t.start_day
                if self.undo_manager:
                    from src.core.undo_manager import MoveTaskCommand
                    self.undo_manager.execute(
                        MoveTaskCommand(self.db, task_id, old_day, new_day)
                    )
                else:
                    t.start_day = new_day
                    self.db.update_task(t)
                self._apply_filter()
                break

    def _on_task_resized(self, task_id: int, new_duration: int):
        """处理甘特条右边缘拖拽调整工期"""
        for t in self.tasks:
            if t.id == task_id:
                old_duration = t.duration
                if self.undo_manager:
                    self.undo_manager.execute(
                        UpdateTaskCommand(self.db, task_id, "duration", old_duration, new_duration)
                    )
                else:
                    t.duration = new_duration
                    self.db.update_task(t)
                self._apply_filter()
                break

    def _on_task_double_clicked(self, task_id: int):
        """双击打开任务编辑器"""
        from src.widgets.task_editor import TaskEditor
        result = TaskEditor.edit(self.db, task_id, self)
        if result:
            self.refresh()

    def _show_context_menu(self, task_id: int, pos: QPoint):
        """右键菜单（支持多选批量操作）"""
        # 获取当前表格中选中行的任务（跳过组头行）
        selected_rows = set(item.row() for item in self.task_table.selectedItems())
        # 确保触发右键的任务行也在选中集合中
        clicked_table_row = -1
        for r, t in enumerate(self._row_to_task):
            if t is not None and t.id == task_id:
                clicked_table_row = r
                break
        if clicked_table_row < 0:
            return
        selected_rows.add(clicked_table_row)
        selected_tasks = [
            self._row_to_task[r]
            for r in sorted(selected_rows)
            if r < len(self._row_to_task) and self._row_to_task[r] is not None
        ]

        menu = QMenu(self)

        if len(selected_tasks) >= 2:
            # ── 批量操作子菜单 ──
            batch_menu = menu.addMenu("📦 批量操作")
            act_batch_done = batch_menu.addAction("✅ 批量标记完成")
            act_batch_undone = batch_menu.addAction("↩️ 批量标记未完成")
            batch_menu.addSeparator()
            act_batch_section = batch_menu.addAction("📁 批量修改分类")
            act_batch_priority = batch_menu.addAction("⚡ 批量修改优先级")
            batch_menu.addSeparator()
            act_batch_delete = batch_menu.addAction("🗑️ 批量删除")

            menu.addSeparator()
            act_add = menu.addAction("➕ 添加任务")

            chosen = menu.exec(pos)
            if chosen == act_batch_done:
                for t in selected_tasks:
                    t.done = True
                    t.progress = 100.0
                    self.db.update_task(t)
                self.refresh()
            elif chosen == act_batch_undone:
                for t in selected_tasks:
                    t.done = False
                    self.db.update_task(t)
                self.refresh()
            elif chosen == act_batch_delete:
                reply = QMessageBox.question(
                    self, "确认批量删除",
                    f"确定要删除选中的 {len(selected_tasks)} 个任务吗？",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
                )
                if reply == QMessageBox.Yes:
                    for t in selected_tasks:
                        self.db.delete_task(t.id)
                    self.refresh()
            elif chosen == act_batch_section:
                all_sections = self.db.get_all_sections()
                if all_sections:
                    names = [s["label"] for s in all_sections]
                    name, ok = QInputDialog.getItem(
                        self, "批量修改分类", "选择新分类:", names, 0, False,
                    )
                    if ok and name:
                        section = next((s for s in all_sections if s["label"] == name), None)
                        if section:
                            for t in selected_tasks:
                                t.section = Section(section["key"])
                                self.db.update_task(t)
                            self.refresh()
            elif chosen == act_batch_priority:
                priorities = ["高", "中", "低"]
                prio, ok = QInputDialog.getItem(
                    self, "批量修改优先级", "选择优先级:", priorities, 1, False,
                )
                if ok:
                    for t in selected_tasks:
                        t.priority = prio
                        self.db.update_task(t)
                    self.refresh()
            elif chosen == act_add:
                self._add_new_task()
        else:
            # ── 单任务操作菜单 ──
            task = selected_tasks[0] if selected_tasks else next(
                (t for t in self.tasks if t.id == task_id), None
            )
            if not task:
                return

            act_edit = menu.addAction("✏️ 编辑任务")
            act_toggle = menu.addAction(
                "↩️ 标记未完成" if task.done else "✅ 标记完成"
            )
            act_progress = menu.addAction("📊 快速调整进度")
            act_duplicate = menu.addAction("📋 复制任务")
            act_delete = menu.addAction("🗑️ 删除任务")
            menu.addSeparator()

            # 测试结果 & Issue 追踪
            act_result = menu.addAction("📝 测试结果记录")
            act_new_issue = menu.addAction("🐛 新建 Issue")
            act_view_issues = menu.addAction("📋 查看 Issue")
            issue_count = self.db.conn.execute(
                "SELECT COUNT(*) FROM test_issues WHERE task_id=?", (task.id,)
            ).fetchone()[0]
            if issue_count > 0:
                act_view_issues.setText(f"📋 查看 Issue ({issue_count})")

            menu.addSeparator()
            act_add = menu.addAction("➕ 添加任务")

            chosen = menu.exec(pos)
            if chosen == act_edit:
                self._on_task_double_clicked(task.id)
            elif chosen == act_toggle:
                task.done = not task.done
                task.progress = 100.0 if task.done else task.progress
                self.db.update_task(task)
                self.refresh()
            elif chosen == act_progress:
                self._quick_progress(task.id)
            elif chosen == act_duplicate:
                self._duplicate_task(task.id)
            elif chosen == act_delete:
                reply = QMessageBox.question(
                    self, "确认删除",
                    f"确定要删除任务 {task.num} {task.name_cn} 吗？",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
                )
                if reply == QMessageBox.Yes:
                    self.db.delete_task(task.id)
                    self.refresh()
            elif chosen == act_add:
                self._add_new_task()
            elif chosen == act_result:
                from src.widgets.test_result_dialog import TestResultDialog
                dlg = TestResultDialog(self.db, task.id, task.num, task.name_cn, self)
                dlg.exec()
            elif chosen == act_new_issue:
                from src.widgets.issue_tracker import _IssueEditDialog
                dlg = _IssueEditDialog(self.db, default_task_id=task.id, parent=self)
                if dlg.exec() == QDialog.DialogCode.Accepted:
                    data = dlg.get_data()
                    self.db.insert_issue(task_id=task.id, title=data["title"],
                                         description=data.get("description", ""),
                                         severity=data.get("severity", "minor"),
                                         priority=data.get("priority", 3),
                                         assignee=data.get("assignee", ""),
                                         tags=data.get("tags", []),
                                         issue_type=data.get("issue_type", "bug"),
                                         phase=data.get("phase", ""))
                    self.refresh()
            elif chosen == act_view_issues:
                main_win = self.window()
                if hasattr(main_win, 'tabs') and hasattr(main_win, 'issue_tracker'):
                    main_win.issue_tracker.filter_by_task(task.id)
                    main_win.tabs.setCurrentIndex(4)

    # ── 搜索/过滤工具栏 ────────────────────────────

    def _setup_toolbar(self, parent_layout: QVBoxLayout):
        """构建搜索/过滤工具栏"""
        toolbar = QWidget()
        toolbar.setStyleSheet("background: #1e1e2e;")
        hbox = QHBoxLayout(toolbar)
        hbox.setContentsMargins(12, 6, 12, 6)
        hbox.setSpacing(8)

        # 项目起始日期
        self._date_label = QLabel(f"📅 {self._project_start_date.isoformat()}")
        self._date_label.setCursor(Qt.CursorShape.PointingHandCursor)
        self._date_label.setStyleSheet("color: #89b4fa; font-size: 13px; padding: 2px 6px;")
        self._date_label.setToolTip("点击修改项目起始日期")
        self._date_label.mousePressEvent = self._on_date_label_clicked
        hbox.addWidget(self._date_label)

        # 搜索框
        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("🔍 搜索编号/名称...")
        self._search_edit.setMaxLength(60)
        self._search_edit.setFixedWidth(250)
        self._search_edit.setStyleSheet(
            "QLineEdit {"
            "  background: #181825; border: 1px solid #313244; border-radius: 4px;"
            "  padding: 4px 8px; color: #cdd6f4; font-size: 13px;"
            "}"
            "QLineEdit:focus { border: 1px solid #89b4fa; }"
        )
        self._search_edit.textChanged.connect(self._apply_filter)
        hbox.addWidget(self._search_edit)

        # 分类过滤下拉框
        self._section_combo = QComboBox()
        self._section_combo.setFixedWidth(150)
        self._section_combo.setStyleSheet(
            "QComboBox {"
            "  background: #181825; border: 1px solid #313244; border-radius: 4px;"
            "  padding: 4px 8px; color: #cdd6f4; font-size: 13px;"
            "}"
            "QComboBox::drop-down { border: none; }"
            "QComboBox QAbstractItemView {"
            "  background: #1e1e2e; color: #cdd6f4; selection-background-color: #313244;"
            "}"
        )
        self._section_combo.addItem("全部分类", "")
        self._refresh_section_combo()
        self._section_combo.currentIndexChanged.connect(self._apply_filter)
        hbox.addWidget(self._section_combo)

        # 过滤计数标签
        self._filter_label = QLabel("显示 0/0 项")
        self._filter_label.setStyleSheet("color: #a6adc8; font-size: 13px;")
        hbox.addWidget(self._filter_label)

        # 分组折叠切换按钮
        self._group_toggle_btn = QPushButton("📂 分组折叠")
        self._group_toggle_btn.setCheckable(True)
        self._group_toggle_btn.setChecked(True)
        self._group_toggle_btn.setFixedWidth(110)
        self._group_toggle_btn.setStyleSheet(
            "QPushButton {"
            "  background: #181825; border: 1px solid #313244; border-radius: 4px;"
            "  padding: 4px 8px; color: #cdd6f4; font-size: 13px;"
            "}"
            "QPushButton:hover { border: 1px solid #89b4fa; }"
            "QPushButton:checked { background: #313244; border: 1px solid #89b4fa; }"
        )
        self._group_toggle_btn.toggled.connect(self._on_group_toggle)
        hbox.addWidget(self._group_toggle_btn)

        hbox.addStretch()
        parent_layout.addWidget(toolbar)

        # ── 统计概览面板 ──
        from src.widgets.stats_panel import StatsPanel
        self.stats_panel = StatsPanel()
        parent_layout.addWidget(self.stats_panel)

    def _refresh_section_combo(self):
        """刷新分类过滤下拉框（保持当前选中）"""
        current_key = self._section_combo.currentData()
        self._section_combo.blockSignals(True)
        self._section_combo.clear()
        self._section_combo.addItem("全部分类", "")
        sections = self.db.get_all_sections()
        for s in sections:
            self._section_combo.addItem(s["label"], s["key"])
        # 恢复之前的选中
        idx = self._section_combo.findData(current_key)
        self._section_combo.setCurrentIndex(max(idx, 0))
        self._section_combo.blockSignals(False)

    def _apply_filter(self):
        """根据搜索关键词和分类过滤任务列表"""
        keyword = self._search_edit.text().strip().lower()
        section_key = self._section_combo.currentData() or ""

        filtered: list[Task] = []
        for t in self.tasks:
            # 分类过滤
            if section_key:
                t_section = t.section.value if isinstance(t.section, Section) else t.section
                if t_section != section_key:
                    continue
            # 关键词过滤（匹配编号或名称）
            if keyword:
                if keyword not in t.num.lower() and keyword not in t.name_cn.lower() and keyword not in t.name_en.lower():
                    continue
            filtered.append(t)

        self._filtered_tasks = filtered
        self._filter_label.setText(f"显示 {len(filtered)}/{len(self.tasks)} 项")
        self._populate_table()
        self._update_gantt()

    # ── 水平滚动同步 ────────────────────────────

    def _sync_timeline_scroll(self, value: int):
        """甘特图水平滚动时同步资源时间线"""
        self.timeline_scroll.horizontalScrollBar().setValue(value)

    # ── 缩放处理 ──────────────────────────────────

    def _on_zoom_changed(self, day_width: int):
        """缩放甘特图时同步资源时间线的单元格宽度"""
        self.resource_timeline._cell_width = day_width
        self.resource_timeline.set_data(
            self._last_schedule_result.get("timeline", {}) if self._last_schedule_result else {},
            self.db.get_all_resources(),
            self._last_schedule_result.get("report", {}).get("total_days", 30) if self._last_schedule_result else 30,
        )

    # ── 键盘快捷键处理 ────────────────────────────

    def _get_selected_task_ids(self) -> list[int]:
        """获取当前表格中选中行的任务 ID（跳过组头行）"""
        ids = []
        for row in range(self.task_table.rowCount()):
            if row >= len(self._row_to_task):
                break
            task = self._row_to_task[row]
            if task is None:
                continue  # 组头行，跳过
            item = self.task_table.item(row, 0)
            if item and self.task_table.isRowSelected(row):
                ids.append(item.data(Qt.ItemDataRole.UserRole))
        return ids

    def _on_delete_selected(self):
        """Delete 快捷键：删除选中任务"""
        ids = self._get_selected_task_ids()
        if not ids:
            return
        if len(ids) == 1:
            task = next((t for t in self.tasks if t.id == ids[0]), None)
            name = f"{task.num} {task.name_cn}" if task else str(ids[0])
            reply = QMessageBox.question(
                self, "确认删除", f"确定要删除任务 {name} 吗？",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
            )
        else:
            reply = QMessageBox.question(
                self, "批量删除", f"确定要删除选中的 {len(ids)} 个任务吗？",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
            )
        if reply == QMessageBox.Yes:
            for tid in ids:
                self.db.delete_task(tid)
            self.refresh()

    def _on_toggle_done(self):
        """Space 快捷键：切换选中任务完成状态"""
        ids = self._get_selected_task_ids()
        if not ids:
            return
        for tid in ids:
            task = next((t for t in self.tasks if t.id == tid), None)
            if task:
                task.done = not task.done
                task.progress = 100.0 if task.done else 0.0
                self.db.update_task(task)
        self.refresh()

    # ── 复制任务 ──────────────────────────────────

    def _duplicate_task(self, task_id: int):
        """复制指定任务，创建副本"""
        import copy
        original = next((t for t in self.tasks if t.id == task_id), None)
        if not original:
            return
        new_task = copy.copy(original)
        new_task.id = 0
        new_task.num = f"{original.num}-副本"
        new_task.done = False
        new_task.progress = 0.0
        new_task.dependencies = []
        new_task.requirements = list(original.requirements)
        from src.widgets.task_editor import TaskEditor
        dlg = TaskEditor(self.db, task=new_task, parent=self)
        dlg.setWindowTitle(f"📋 复制任务: {original.num}")
        if dlg.exec() == QDialog.DialogCode.Accepted:
            task = dlg.get_task()
            task.id = self.db.insert_task(task)
            self.refresh()

    # ── 进度快速调整 ──────────────────────────────

    def _quick_progress(self, task_id: int):
        """弹出小对话框快速调整任务进度"""
        task = next((t for t in self.tasks if t.id == task_id), None)
        if not task:
            return

        dlg = QDialog(self)
        dlg.setWindowTitle(f"📊 调整进度 - {task.num} {task.name_cn}")
        dlg.setMinimumSize(360, 140)
        dlg.resize(360, 140)
        dlg.setStyleSheet(
            "QDialog { background: #1e1e2e; }"
            "QLabel { color: #cdd6f4; font-size: 14px; }"
            "QSlider::groove:horizontal {"
            "  background: #313244; height: 8px; border-radius: 4px;"
            "}"
            "QSlider::handle:horizontal {"
            "  background: #89b4fa; width: 18px; margin: -5px 0; border-radius: 9px;"
            "}"
            "QSlider::sub-page:horizontal { background: #a6e3a1; border-radius: 4px; }"
            "QPushButton {"
            "  background: #181825; border: 1px solid #313244; border-radius: 4px;"
            "  padding: 6px 16px; color: #cdd6f4; font-size: 13px;"
            "}"
            "QPushButton:hover { border: 1px solid #89b4fa; }"
        )

        layout = QVBoxLayout(dlg)
        layout.setSpacing(10)

        value_label = QLabel(f"当前进度: {task.progress:.0f}%")
        value_label.setAlignment(Qt.AlignCenter)

        slider = QSlider(Qt.Horizontal)
        slider.setRange(0, 100)
        slider.setValue(int(task.progress))
        slider.valueChanged.connect(lambda v: value_label.setText(f"当前进度: {v}%"))

        btn_layout = QHBoxLayout()
        btn_cancel = QPushButton("取消")
        btn_ok = QPushButton("确认")
        btn_ok.setStyleSheet(btn_ok.styleSheet() + "QPushButton { background: #181825; }")

        layout.addWidget(value_label)
        layout.addWidget(slider)
        layout.addLayout(btn_layout)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_ok)

        btn_cancel.clicked.connect(dlg.reject)
        btn_ok.clicked.connect(dlg.accept)

        if dlg.exec() == QDialog.DialogCode.Accepted:
            old_progress = task.progress
            new_progress = float(slider.value())
            old_done = task.done
            new_done = new_progress >= 100.0

            if self.undo_manager:
                from src.core.undo_manager import UpdateProgressCommand
                self.undo_manager.execute(
                    UpdateProgressCommand(self.db, task.id, int(old_progress), int(new_progress))
                )
                # Also update the done field via direct DB call
                self.db.update_task_fields(task.id, {"done": new_done})
            else:
                task.progress = new_progress
                task.done = new_done
                self.db.update_task(task)
            self.refresh()

    def _add_new_task(self):
        """添加新任务"""
        from src.widgets.task_editor import TaskEditor
        result = TaskEditor.add_new(self.db, self)
        if result:
            self.refresh()

    def show_scheduler_dialog(self):
        """显示自动排程对话框"""
        from src.widgets.scheduler_dialog import SchedulerDialog
        if self._scheduler_dialog is None:
            self._scheduler_dialog = SchedulerDialog(self.db, self)
        if self._scheduler_dialog.exec() == 1:
            # 排程完成后刷新并更新时间线
            self.refresh()
            # 重新跑排程获取最新时间线
            resources = self.db.get_all_resources()
            config = ScheduleConfig(mode=ScheduleMode.BALANCED)
            result = run_auto_schedule(self.tasks, resources, config)
            self._last_schedule_result = result
            self._update_timeline()

    def export_to_excel(self, path: str):
        """导出到 Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "可靠性测试排程"

        headers = ["编号", "测试项目(中文)", "测试项目(英文)", "分类",
                   "天数", "开始日", "结束日", "样品池", "样品数",
                   "优先级", "状态"]
        header_fill = PatternFill(start_color="313244", end_color="313244", fill_type="solid")
        header_font = Font(color="CDD6F4", bold=True, size=11)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 动态构建 section 填充色
        section_fills = {}
        colors_map = getattr(self, '_section_colors', None)
        for s in self.db.get_all_sections():
            hex_color = s["color"].lstrip("#")
            section_fills[s["key"]] = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )
        thin_border = Border(
            left=Side(style="thin", color="313244"),
            right=Side(style="thin", color="313244"),
            top=Side(style="thin", color="313244"),
            bottom=Side(style="thin", color="313244"),
        )

        for i, t in enumerate(self.tasks, 2):
            data = [
                t.num, t.name_cn, t.name_en,
                get_section_label(t.section, getattr(self, '_section_labels', None)),
                t.duration, f"D{t.start_day + 1}", f"D{t.start_day + t.duration}",
                t.sample_pool, t.sample_qty,
                t.priority, "✅ 完成" if t.done else "⏳ 进行中",
            ]
            section_key = t.section.value if isinstance(t.section, Section) else t.section
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(color="1e1e2e" if section_fills.get(section_key) else "CDD6F4")

            fill = section_fills.get(section_key)
            if fill:
                ws.cell(row=i, column=4).fill = fill

        col_widths = [8, 20, 25, 12, 6, 8, 8, 10, 8, 8, 10]
        for i, cw in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = cw

        wb.save(path)
