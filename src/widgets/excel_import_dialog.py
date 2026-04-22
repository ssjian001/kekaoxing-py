"""Excel 批量导入对话框 — 从 .xlsx 文件批量导入测试项目到数据库。"""

from __future__ import annotations

import logging
import sys
from pathlib import Path
from typing import Optional

from PySide6.QtCore import Qt, QTimer
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QRadioButton, QButtonGroup, QTableWidget, QTableWidgetItem,
    QHeaderView, QFileDialog, QMessageBox, QSizePolicy,
    QProgressDialog,
)

from src.models import Section, DEFAULT_SECTION_LABELS

# ── Catppuccin Mocha 样式 ──────────────────────────────────────────────
CATPPUCCIN_QSS = """
QDialog {
    background-color: #1e1e2e;
    color: #cdd6f4;
}
QTableWidget {
    background-color: #1e1e2e;
    alternate-background-color: #181825;
    color: #cdd6f4;
    gridline-color: #313244;
    border: 1px solid #313244;
    selection-background-color: #45475a;
    font-size: 13px;
}
QTableWidget::item {
    padding: 4px 8px;
}
QTableWidget::item:selected {
    background-color: #45475a;
}
QHeaderView::section {
    background-color: #181825;
    color: #bac2de;
    border: none;
    border-bottom: 1px solid #313244;
    border-right: 1px solid #313244;
    padding: 6px 8px;
    font-weight: bold;
    font-size: 13px;
}
QPushButton {
    background-color: #313244;
    color: #cdd6f4;
    border: 1px solid #45475a;
    border-radius: 4px;
    padding: 6px 14px;
    font-size: 13px;
    min-width: 70px;
}
QPushButton:hover {
    background-color: #45475a;
}
QPushButton:pressed {
    background-color: #585b70;
}
QPushButton:disabled {
    color: #585b70;
    border-color: #313244;
}
QPushButton#btnImport {
    background-color: #1e6640;
    color: #a6e3a1;
    border-color: #2d9f5f;
    font-weight: bold;
}
QPushButton#btnImport:hover {
    background-color: #2d9f5f;
}
QPushButton#btnImport:disabled {
    background-color: #313244;
    color: #585b70;
    border-color: #45475a;
}
QPushButton#btnCancel {
    background-color: #6e2535;
    color: #f38ba8;
    border-color: #a6374d;
}
QPushButton#btnCancel:hover {
    background-color: #a6374d;
}
QLabel {
    color: #cdd6f4;
    font-size: 13px;
}
QLabel#fileLabel {
    color: #a6adc8;
    font-style: italic;
}
QLabel#titleLabel {
    color: #cdd6f4;
    font-size: 15px;
    font-weight: bold;
}
QLabel#mappingLabel {
    color: #a6adc8;
    font-size: 12px;
}
QRadioButton {
    color: #cdd6f4;
    font-size: 13px;
    spacing: 6px;
}
QRadioButton::indicator {
    width: 14px;
    height: 14px;
    border: 1px solid #45475a;
    border-radius: 7px;
    background-color: #313244;
}
QRadioButton::indicator:checked {
    background-color: #89b4fa;
    border-color: #89b4fa;
}
QScrollBar:vertical {
    background: #1e1e2e;
    width: 10px;
    border: none;
}
QScrollBar::handle:vertical {
    background: #45475a;
    border-radius: 5px;
    min-height: 30px;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0;
}
QScrollBar:horizontal {
    background: #1e1e2e;
    height: 10px;
    border: none;
}
QScrollBar::handle:horizontal {
    background: #45475a;
    border-radius: 5px;
    min-width: 30px;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    width: 0;
}
"""

# ── 列匹配规则 ──────────────────────────────────────────────────────────
# Excel 表头文字（不区分大小写）与数据库字段的映射
COLUMN_PATTERNS: dict[str, list[str]] = {
    "num":        ["编号", "num", "序号", "no"],
    "name_cn":    ["中文", "name_cn", "项目(中文)", "测试项目(中文)", "名称"],
    "name_en":    ["英文", "name_en", "项目(英文)", "测试项目(英文)", "english"],
    "section":    ["分类", "section", "类别", "类型"],
    "duration":   ["天数", "duration", "工时", "周期"],
    "start_day":  ["开始", "start", "开始日", "开始天"],
    "sample_pool":["样品池", "sample_pool", "pool"],
    "sample_qty": ["样品数", "sample_qty", "数量"],
    "priority":   ["优先级", "priority", "权重"],
}

# section 中文标签 → key 的反向映射（内置默认 + 运行时动态更新）
SECTION_REVERSE_MAP: dict[str, str] = {v: k for k, v in DEFAULT_SECTION_LABELS.items()}


def _match_column(header: str) -> Optional[str]:
    """用表头文字模糊匹配字段名，返回字段 key 或 None。"""
    h = header.strip().lower()
    for field_key, patterns in COLUMN_PATTERNS.items():
        if any(pat.lower() in h for pat in patterns):
            return field_key
    return None


def _resolve_section(value) -> Optional[str]:
    """将 section 列的值（可能是中文标签或 key）转为 section key。"""
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None

    # 先尝试直接匹配 key（env / mech / surf / pack）
    for key in Section:
        if text == key.value.lower():
            return key.value

    # 再尝试中文标签反向查找
    for label, key in SECTION_REVERSE_MAP.items():
        if text == label.lower() or label.lower() in text:
            return key

    return None


def _parse_day_value(value) -> Optional[int]:
    """解析天数/开始日，支持纯数字或 'D1' 格式。"""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.upper().startswith("D"):
        text = text[1:]
    try:
        return int(float(text))
    except (ValueError, TypeError):
        return None


class ExcelImportDialog(QDialog):
    """Excel 批量导入对话框。

    Usage::

        dlg = ExcelImportDialog(db, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            print(f"导入了 {dlg.imported_count} 条")
    """

    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.imported_count: int = 0
        self.skipped_count: int = 0

        # 解析后的数据
        self._parsed_rows: list[dict] = []
        self._column_mapping: dict[int, str] = {}   # Excel 列索引 → 字段名
        self._all_headers: list[str] = []            # 原始表头

        self.setWindowTitle("📋 导入 Excel")
        self.setMinimumSize(780, 520)
        self.setStyleSheet(CATPPUCCIN_QSS)
        self._build_ui()

    # ── UI 构建 ──────────────────────────────────────────────────────────
    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 12)
        layout.setSpacing(10)

        # 标题
        title = QLabel("📋 从 Excel 导入测试项目")
        title.setObjectName("titleLabel")
        layout.addWidget(title)

        # 文件选择行
        file_row = QHBoxLayout()
        self.btn_file = QPushButton("📁 选择文件")
        self.btn_file.setFixedWidth(120)
        self.btn_file.clicked.connect(self._on_select_file)
        file_row.addWidget(self.btn_file)

        self.file_label = QLabel("未选择文件")
        self.file_label.setObjectName("fileLabel")
        self.file_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        file_row.addWidget(self.file_label)
        layout.addLayout(file_row)

        # 列映射信息
        self.mapping_label = QLabel("")
        self.mapping_label.setObjectName("mappingLabel")
        self.mapping_label.setWordWrap(True)
        self.mapping_label.setTextFormat(Qt.TextFormat.RichText)
        layout.addWidget(self.mapping_label)

        # 预览表格
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        layout.addWidget(self.table, stretch=1)

        # 底部：导入模式 + 按钮
        bottom = QHBoxLayout()

        self.radio_append = QRadioButton("追加（默认）")
        self.radio_append.setChecked(True)
        self.radio_overwrite = QRadioButton("覆盖全部")
        mode_group = QButtonGroup(self)
        mode_group.addButton(self.radio_append)
        mode_group.addButton(self.radio_overwrite)
        bottom.addWidget(self.radio_append)
        bottom.addWidget(self.radio_overwrite)
        bottom.addStretch()

        self.btn_import = QPushButton("✅ 导入")
        self.btn_import.setObjectName("btnImport")
        self.btn_import.setEnabled(False)
        self.btn_import.clicked.connect(self._on_import)

        self.btn_cancel = QPushButton("❌ 取消")
        self.btn_cancel.setObjectName("btnCancel")
        self.btn_cancel.clicked.connect(self.reject)

        bottom.addWidget(self.btn_import)
        bottom.addWidget(self.btn_cancel)
        layout.addLayout(bottom)

    # ── 文件选择 ─────────────────────────────────────────────────────────
    def _on_select_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "",
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if not path:
            return

        try:
            self._parse_excel(path)
        except Exception as e:
            QMessageBox.critical(self, "解析失败", f"无法解析 Excel 文件：\n{e}")
            return

        self.file_label.setText(Path(path).name)
        self.file_label.setObjectName("fileLabel")
        self.file_label.setStyleSheet("color: #a6e3a1; font-style: normal;")
        self._refresh_mapping_info()
        self._refresh_preview_table()
        self.btn_import.setEnabled(True)

    # ── Excel 解析 ───────────────────────────────────────────────────────
    def _parse_excel(self, path: str):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        # 第一行 = 表头
        try:
            headers = next(rows_iter)
        except StopIteration:
            raise ValueError("Excel 文件为空，没有表头行。")

        self._all_headers = [str(h) if h is not None else "" for h in headers]

        # 智能列匹配
        self._column_mapping = {}
        for col_idx, header in enumerate(self._all_headers):
            field = _match_column(header)
            if field is not None:
                self._column_mapping[col_idx] = field

        # 更新 section 反向映射（从数据库读取自定义分类）
        self._update_section_reverse_map()

        # 解析数据行
        self._parsed_rows = []
        for row in rows_iter:
            if all(c is None for c in row):
                continue
            row_dict: dict = {}
            for col_idx, field in self._column_mapping.items():
                val = row[col_idx] if col_idx < len(row) else None
                row_dict[field] = val
            self._parsed_rows.append(row_dict)

        wb.close()

    def _update_section_reverse_map(self):
        """从数据库读取所有分类（包括自定义分类），更新反向映射。"""
        global SECTION_REVERSE_MAP
        SECTION_REVERSE_MAP = {v: k for k, v in DEFAULT_SECTION_LABELS.items()}
        try:
            sections = self.db.get_all_sections()
            for s in sections:
                if s.get("key") and s.get("label"):
                    SECTION_REVERSE_MAP[s["label"]] = s["key"]
        except Exception:
            logging.debug("更新section反向映射失败", exc_info=True)

    # ── 映射信息 & 预览 ─────────────────────────────────────────────────
    def _refresh_mapping_info(self):
        if not self._all_headers:
            self.mapping_label.setText("")
            return

        lines = []
        for col_idx, header in enumerate(self._all_headers):
            field = self._column_mapping.get(col_idx)
            if field:
                lines.append(
                    f'<span style="color:#a6e3a1;">✓</span> '
                    f'<b>{header}</b> → <span style="color:#89b4fa;">{field}</span>'
                )
            else:
                lines.append(
                    f'<span style="color:#585b70;">✗</span> '
                    f'<span style="color:#6c7086;">{header}</span> '
                    f'<span style="color:#585b70;">（忽略）</span>'
                )

        matched = len(self._column_mapping)
        total_cols = len(self._all_headers)
        summary = f"已匹配 <b>{matched}</b>/{total_cols} 列，共 <b>{len(self._parsed_rows)}</b> 行数据"
        self.mapping_label.setText(f"{summary}<br>{' &nbsp;│&nbsp; '.join(lines)}")

    def _refresh_preview_table(self):
        # 显示列：只显示匹配到的字段 + 未匹配但原始表头中有值的列
        display_fields = [
            (col_idx, self._column_mapping[col_idx])
            for col_idx in sorted(self._column_mapping.keys())
        ]
        # 加上未匹配的列
        for col_idx, header in enumerate(self._all_headers):
            if col_idx not in self._column_mapping:
                display_fields.append((col_idx, header))

        headers = [self._all_headers[ci] for ci, _ in display_fields]
        preview_rows = self._parsed_rows[:20]

        self.table.setColumnCount(len(headers))
        self.table.setRowCount(len(preview_rows))
        self.table.setHorizontalHeaderLabels(headers)

        for r, row_data in enumerate(preview_rows):
            for c, (col_idx, _) in enumerate(display_fields):
                val = row_data.get(self._column_mapping.get(col_idx, col_idx), "")
                item = QTableWidgetItem(str(val) if val is not None else "")
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                # 未匹配列灰显
                if col_idx not in self._column_mapping:
                    item.setForeground(Qt.GlobalColor.gray)
                self.table.setItem(r, c, item)

        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )

    # ── 导入执行 ─────────────────────────────────────────────────────────
    def _on_import(self):
        if not self._parsed_rows:
            QMessageBox.warning(self, "无数据", "没有可导入的数据行。")
            return

        overwrite = self.radio_overwrite.isChecked()

        if overwrite:
            confirm = QMessageBox.warning(
                self, "确认覆盖",
                "⚠️ 覆盖模式将删除数据库中所有现有测试项目！\n\n确定继续？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if confirm != QMessageBox.StandardButton.Yes:
                return
            self.db.conn.execute("DELETE FROM tasks")
            # 级联清理关联数据（防止外键 CASCADE 未生效时产生孤儿记录）
            self.db.conn.execute("DELETE FROM issue_history")
            self.db.conn.execute("DELETE FROM test_issues")
            self.db.conn.execute("DELETE FROM test_results")

        imported = 0
        skipped = 0

        progress = QProgressDialog("正在导入数据...", None, 0, len(self._parsed_rows), self)
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(0)
        progress.setCancelButton(None)

        try:
            for i, row_data in enumerate(self._parsed_rows):
                progress.setValue(i)
                # 至少需要 name_cn 或 name_en 之一才算有效行
                name_cn = str(row_data.get("name_cn", "") or "").strip()
                name_en = str(row_data.get("name_en", "") or "").strip()
                if not name_cn and not name_en:
                    skipped += 1
                    continue

                # 解析 section
                raw_section = row_data.get("section")
                section_val = _resolve_section(raw_section)
                if section_val is None:
                    section_val = Section.ENV.value

                # 解析数值字段
                duration = _parse_day_value(row_data.get("duration")) or 1
                start_day = _parse_day_value(row_data.get("start_day")) or 0
                sample_qty = _parse_day_value(row_data.get("sample_qty")) or 3
                priority = _parse_day_value(row_data.get("priority")) or 3

                task_dict = {
                    "num": str(row_data.get("num", i + 1) or f"{i + 1}"),
                    "name_cn": name_cn,
                    "name_en": name_en,
                    "section": section_val,
                    "duration": duration,
                    "start_day": start_day,
                    "sample_pool": str(row_data.get("sample_pool", "product") or "product").strip(),
                    "sample_qty": sample_qty,
                    "priority": priority,
                }

                try:
                    self.db.insert_task_dict(task_dict)
                    imported += 1
                except Exception:
                    skipped += 1
                    logging.debug("导入任务行失败: num=%s, data=%s", task_dict.get("num"), task_dict, exc_info=True)
        finally:
            progress.close()

        self.db.conn.commit()
        self.imported_count = imported
        self.skipped_count = skipped

        # 结果提示
        mode_text = "覆盖导入" if overwrite else "追加导入"
        msg = (
            f"✅ {mode_text}完成！\n\n"
            f"成功导入：<b>{imported}</b> 条\n"
            f"跳过（数据不完整）：<b>{skipped}</b> 条"
        )
        QMessageBox.information(self, "导入完成", msg)
        self.accept()
