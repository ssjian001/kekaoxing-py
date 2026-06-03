"""导出服务 — 共享工具模块。

跨平台 CJK 字体检测、Excel 样式构造、路径校验、结论判定等。
"""

from __future__ import annotations

import json
import logging
import os
import platform
import subprocess
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

from src.constants import (
    TASK_STATUS_LABELS,
    ISSUE_STATUS_LABELS,
    SAMPLE_STATUS_LABELS,
)

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


# ── 类别中文映射 ─────────────────────────────────────────────

CATEGORY_MAP = {
    "环境试验": "env", "env": "环境试验",
    "机械试验": "mech", "mech": "机械试验",
    "表面处理": "surf", "surf": "表面处理",
    "包装": "pack", "pack": "包装",
    "其他": "other", "other": "其他",
}

# Merged status → Chinese label map
STATUS_MAP: dict[str, str] = {
    **TASK_STATUS_LABELS,
    **ISSUE_STATUS_LABELS,
    **SAMPLE_STATUS_LABELS,
}


# ── CJK 字体 ─────────────────────────────────────────────────

_CJK_FONT: str | None = None


def get_cjk_font() -> str:
    """返回当前平台可用的中文字体名（单个字体，非 fallback 链）。

    优先级：Microsoft YaHei > PingFang SC > Noto Sans CJK SC > sans-serif。
    缓存到模块变量，只检测一次。
    """
    global _CJK_FONT
    if _CJK_FONT is not None:
        return _CJK_FONT

    system = platform.system()
    if system == "Windows":
        _CJK_FONT = "Microsoft YaHei"
    elif system == "Darwin":
        _CJK_FONT = "PingFang SC"
    else:
        for candidate in ("Noto Sans CJK SC", "WenQuanYi Micro Hei", "Droid Sans Fallback"):
            try:
                r = subprocess.run(
                    ["fc-match", "-f", "%{family}", candidate],
                    capture_output=True, text=True, timeout=3,
                )
                if r.returncode == 0 and r.stdout.strip():
                    _CJK_FONT = r.stdout.strip()
                    break
            except Exception:
                continue
        if _CJK_FONT is None:
            _CJK_FONT = "Noto Sans CJK SC"
    return _CJK_FONT


# ── Excel 共享样式 ────────────────────────────────────────────

def excel_styles(fill_color: str = "2B579A") -> dict:
    """返回 Excel 通用样式字典（字体、填充、对齐、边框）。"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    _f = get_cjk_font()
    return {
        "header_font": Font(name=_f, size=11, bold=True, color="FFFFFF"),
        "header_fill": PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid"),
        "cell_font": Font(name=_f, size=10),
        "title_font": lambda c: Font(name=_f, size=14, bold=True, color=c),
        "sub_font": Font(name=_f, size=9, color="666666"),
        "center": Alignment(horizontal="center", vertical="center"),
        "thin_border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
    }


def excel_write_title_block(ws, title: str, merge_range: str, subtitle: str, sub_range: str,
                             title_color: str, styles: dict) -> None:
    """写入 Excel 标题行 + 副标题行。"""
    from openpyxl.styles import Alignment
    ws.merge_cells(merge_range)
    title_cell = ws[merge_range.split(":")[0]]
    title_cell.value = title
    title_cell.font = styles["title_font"](title_color)
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(sub_range)
    sub = ws[sub_range.split(":")[0]]
    sub.value = subtitle
    sub.font = styles["sub_font"]
    sub.alignment = Alignment(horizontal="center")


def excel_write_headers(ws, row: int, headers: list[str], styles: dict) -> None:
    """写入 Excel 表头行。"""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["thin_border"]


def excel_write_row(ws, row: int, values: list, styles: dict, alignment=None) -> None:
    """写入 Excel 单行数据。"""
    align = alignment or styles["center"]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = styles["cell_font"]
        cell.alignment = align
        cell.border = styles["thin_border"]


def _sanitize_filename(name: str) -> str:
    """移除文件名非法字符，防止路径分隔符混入文件名。"""
    import re
    # Windows 非法字符：\ / : * ? " < > |
    return re.sub(r'[\\/:*?"<>|]', '_', name)


def _sanitize_path(path: str) -> str:
    """清理完整路径中的非法文件名字符（只处理末尾的文件名部分）。"""
    import re
    # 只替换路径末尾文件名部分的非法字符，不动目录分隔符
    safe = re.sub(r'[\\:*?"<>|]', '_', path)
    # 路径中的正斜杠保留（Unix/URL 合法），反斜杠替换
    safe = safe.replace('\\', '_')
    return safe


def excel_save(wb, filepath: str | None, filename: str, output_dir: Path) -> str:
    """保存 Excel 工作簿，返回绝对路径。"""
    from openpyxl.utils import get_column_letter

    # 同时清理 filename（自动命名）和 out（用户指定路径）
    safe_name = _sanitize_filename(filename)
    raw_out = filepath or str(output_dir / safe_name)
    safe_out = _sanitize_path(raw_out)
    resolved = _validate_output_path(safe_out, output_dir)
    try:
        wb.save(str(resolved))
    except (OSError, PermissionError) as e:
        logger.error("Excel save failed: %s → %s", resolved, e)
        raise
    return str(resolved)


# ── 路径校验 ──────────────────────────────────────────────────

def _validate_output_path(path: str | Path, output_dir: Path) -> Path:
    """校验输出路径安全性，防止路径遍历。"""
    import tempfile

    resolved = Path(path).resolve()
    try:
        resolved.relative_to(output_dir.resolve())
    except ValueError:
        tmp = Path(tempfile.gettempdir()).resolve()
        if not resolved.is_relative_to(tmp):
            raise ValueError(f"导出路径超出允许范围: {resolved}")
    return resolved


# ── 结论判定 ──────────────────────────────────────────────────

def _judge_conclusion(
    pass_count: int, fail_count: int, conditional_count: int,
    total_results: int, accept_criteria: str = "",
) -> str:
    """基于接收准则判定结论。

    结构化 JSON 格式:
      {"type": "c0"} — 全数通过，fail=0 才 PASS
      {"type": "aql", "accept": N} — fail<=N 则 PASS
      {"type": "custom", "accept": N, "reject": M}

    纯文本或无法解析时退化: fail>0→FAIL, conditional>0→CONDITIONAL, pass>0→PASS
    """
    if accept_criteria:
        try:
            start = accept_criteria.find("{")
            end = accept_criteria.rfind("}") + 1
            if start >= 0 and end > start:
                criteria = json.loads(accept_criteria[start:end])
                ctype = criteria.get("type", "")
                if ctype == "c0":
                    if fail_count > 0:
                        return "FAIL"
                    elif conditional_count > 0:
                        return "CONDITIONAL"
                    elif pass_count > 0:
                        return "PASS"
                    return "—"
                elif ctype in ("aql", "custom"):
                    accept_n = criteria.get("accept", 0)
                    if fail_count > accept_n:
                        return "FAIL"
                    elif conditional_count > 0 and fail_count == accept_n:
                        return "CONDITIONAL"
                    elif pass_count > 0:
                        return "PASS"
                    return "—"
        except (json.JSONDecodeError, ValueError):
            pass
    if fail_count > 0:
        return "FAIL"
    elif conditional_count > 0:
        return "CONDITIONAL"
    elif pass_count > 0:
        return "PASS"
    return "—"