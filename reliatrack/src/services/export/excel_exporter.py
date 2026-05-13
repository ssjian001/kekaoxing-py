"""导出服务 — Excel 导出器。

包含：测试任务导出、Issue 导出、样品台账导出（openpyxl）。
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from src.models.issue import Issue, FARecord, CAPARecord
    from src.models.sample import Sample
    from src.models.test_plan import TestPlan, TestTask, TestResult

from src.services.export.export_utils import (
    CATEGORY_MAP, STATUS_MAP, excel_styles, excel_write_title_block,
    excel_write_headers, excel_write_row, excel_save,
)
from src.constants import RESOLUTION_LABELS

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
else:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter


def export_tasks_excel(
    output_dir: Path,
    plan: TestPlan,
    tasks: list[TestTask],
    results: list[TestResult] | None = None,
    technician_names: dict[int, str] | None = None,
    filepath: str | None = None,
) -> str:
    """导出测试任务列表为 Excel。"""
    s = excel_styles("2B579A")
    wb = Workbook()
    ws = wb.active
    ws.title = "测试任务"

    from datetime import date, timedelta
    plan_start = None
    if plan.start_date:
        try:
            plan_start = date.fromisoformat(plan.start_date)
        except ValueError:
            pass

    res_map: dict[int, tuple[int, int]] = {}
    if results:
        for r in results:
            if r.task_id:
                p, t = res_map.get(r.task_id, (0, 0))
                if r.result == "pass":
                    p += 1
                res_map[r.task_id] = (p, t + 1)

    excel_write_title_block(
        ws, f"测试计划: {plan.name}", "A1:M1",
        f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  测试标准: {plan.test_standard or '—'}",
        "A2:M2", "2B579A", s,
    )

    headers = ["序号", "名称", "类别", "天数", "预计开始", "预计结束", "进度", "优先级", "状态", "技术员", "通过率", "实际开始", "实际完成"]
    excel_write_headers(ws, 4, headers, s)

    prefix = getattr(plan, 'task_prefix', '') or ''
    for seq, task in enumerate(tasks, 1):
        task_id_display = f"{prefix}-{seq:03d}" if prefix else (task.id or seq)
        row_idx = seq + 4
        if plan_start and task.start_day is not None:
            planned_start = (plan_start + timedelta(days=task.start_day)).isoformat()
            planned_end = (plan_start + timedelta(days=task.start_day + task.duration - 1)).isoformat()
        else:
            planned_start = str(task.start_day) if task.start_day else "—"
            planned_end = "—"
        pass_count, total = res_map.get(task.id, (0, 0)) if task.id else (0, 0)
        rate_text = f"{pass_count}/{total}" if total > 0 else "—"
        tech_name = (technician_names or {}).get(task.technician_id, "") if task.technician_id else ""

        values = [
            task_id_display,
            task.name,
            CATEGORY_MAP.get(task.category, task.category),
            task.duration,
            planned_start,
            planned_end,
            f"{task.progress:.0f}%",
            task.priority,
            STATUS_MAP.get(task.status, task.status),
            tech_name,
            rate_text,
            task.actual_start_date or "—",
            task.actual_end_date or "—",
        ]
        excel_write_row(ws, row_idx, values, s)

    widths = [5, 25, 10, 8, 12, 12, 8, 8, 10, 12, 10, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return excel_save(wb, filepath, f"测试任务_{plan.name}_{datetime.now():%Y%m%d_%H%M}.xlsx", output_dir)


def export_issues_excel(
    output_dir: Path,
    issues: list[Issue],
    fa_map: dict[int, list[FARecord]] | None = None,
    capa_map: dict[int, list[CAPARecord]] | None = None,
    filepath: str | None = None,
) -> str:
    """导出 Issue 列表为 Excel。"""
    from openpyxl.styles import Alignment

    s = excel_styles("C0504D")
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Issue 追踪"

    excel_write_title_block(
        ws, "Issue 追踪报告", "A1:K1",
        f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  共 {len(issues)} 个 Issue",
        "A2:K2", "C0504D", s,
    )

    headers = ["ID", "Issue描述", "严重度", "状态", "优先级", "DRI", "报告人", "解决结果", "根因分析", "改善对策", "CAPA 状态"]
    excel_write_headers(ws, 4, headers, s)

    for row_idx, issue in enumerate(issues, 5):
        capas = capa_map.get(issue.id, []) if capa_map and issue.id is not None else []
        capa_actions = "; ".join(c.action for c in capas if c.action) or ""
        capa_statuses = "; ".join(c.status for c in capas) or ""
        values = [
            issue.id,
            issue.title,
            issue.severity,
            STATUS_MAP.get(issue.status, issue.status),
            issue.priority,
            issue.dri_name or "",
            issue.reporter_name or "",
            RESOLUTION_LABELS.get(issue.resolution, issue.resolution) or "",
            (issue.root_cause or "")[:100],
            issue.improvement_measures or "",
            capa_statuses,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = s["cell_font"]
            cell.alignment = wrap if col in (2, 8, 9, 10, 11) else s["center"]
            cell.border = s["thin_border"]

    widths = [5, 25, 10, 10, 8, 12, 12, 12, 35, 35, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return excel_save(wb, filepath, f"Issue追踪_{datetime.now():%Y%m%d_%H%M}.xlsx", output_dir)


def export_samples_excel(
    output_dir: Path,
    samples: list[Sample],
    filepath: str | None = None,
) -> str:
    """导出样品台账为 Excel。"""
    s = excel_styles("4F81BD")

    wb = Workbook()
    ws = wb.active
    ws.title = "样品台账"

    excel_write_title_block(
        ws, "样品台账", "A1:F1",
        f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  共 {len(samples)} 个样品",
        "A2:F2", "4F81BD", s,
    )

    headers = ["ID", "SN", "批次号", "规格型号", "状态", "存放位置"]
    excel_write_headers(ws, 4, headers, s)

    for row_idx, sample in enumerate(samples, 5):
        values = [
            sample.id,
            sample.sn,
            sample.batch_no or "",
            sample.spec or "",
            STATUS_MAP.get(sample.status, sample.status),
            sample.location or "",
        ]
        excel_write_row(ws, row_idx, values, s)

    widths = [5, 20, 15, 20, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return excel_save(wb, filepath, f"样品台账_{datetime.now():%Y%m%d_%H%M}.xlsx", output_dir)